## 1.2.2版本更新日志：

## 目前已实现的功能有：

## 1. 随机抽取指定数量的单词进行考查；
## 2. 具备两种考查模式：单元考查模式（只考查一个List的单词） 和 综合考查模式（可以考查多个List的单词）；
## 3. 自动根据测试时间实时生成错题本；

## 功能上还需要完善的： 1. 添加选择题选项 
##                    2.计算正确率并制图（答题时间和正确率的关系）

import pandas as pd 
import numpy as np 
import random as rm 
import openpyxl 
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def test_initialization_lttype():

	continueTest = True
	while continueTest:
		print("\n")
		test_list = int(input(" \t 你想要考查哪一个list的单词?"))
		print("\n")
		global test_quantity_lttype
		test_quantity_lttype = input(" \t 你想要考查多少单词？")
		if not test_quantity_lttype.isdigit():
			if test_quantity_lttype == "all" or test_quantity_lttype == "ALL":
				test_quantity_lttype = len(gre["List" + " " + str(test_list)])
		else:
			test_quantity_lttype = int(test_quantity_lttype)
			if test_quantity_lttype >= len(gre["List"+" "+str(test_list)]):
				test_quantity_lttype = len(gre["List"+" "+str(test_list)])
		word_list = rm.choices(population = gre["List"+" "+str(test_list)], k = test_quantity_lttype)

		global indices
		indices = dict()
		for i in word_list:
			indices[i] = gre["List"+" "+str(test_list)].index(i)

		global answers
		answers = dict()
		for i in indices.values():
			answers[i] = gre["Answer"+" "+str(test_list)][i]

		global scoreGained
		scoreGained = 0
		global scoreLost
		scoreLost = 0
		global wrongWordsList
		wrongWordsList = dict()
		print("\n")
		print("\t" + "请回答以下各个单词的中文含义是什么:" + "\n")
		for i in indices.keys():
			print("\t" + i + "\t")
			print("\n")
			usr_answer = input("\t" +"中文含义: \t")
			print("\n" + "\t\t\t\t\t\t\t标准答案是：" + str(answers[indices[i]]))
			print("\n")
			usr_scoring = input("回答是否正确（是/否）: ")
			if usr_scoring == "Y" or usr_scoring == "y" or usr_scoring == "是":
				scoreGained += 1
			else:
				scoreLost += 1
				wrongWordsList[i] = answers[indices[i]]
			print("\n")
		print("\t" + "你回答了"+ str(test_quantity_lttype) +"个单词的含义；其中"+ str(scoreGained) +"个正确，"+ str(scoreLost) +"个错误。")
		print("\n")
		if scoreLost > 0:
			print("其中错误的单词为：")
			print("\n")
			for i in wrongWordsList.keys():
				print("\t\t\t\t" + str(i) + ", " + str(wrongWordsList[i]))
				print("\n")
		global test_quantity
		test_quantity = test_quantity_lttype
		review_wrong_words()
		continueTest = input("是否继续？")
		if continueTest == "Y" or continueTest == "y" or continueTest == "是":
			continueTest = True
		else:
			continueTest = False

def test_initialization_wltype():
	
	continueTest = True
	while continueTest:
		print("\n")
		test_list = input(" \t 你想要考查哪几个list的单词（请用英文逗号进行划分）?").strip(',')
		global test_list_range
		test_list_range = list(filter(lambda a: a != ",", test_list))

		gre_subset = dict()
		for i in test_list_range:
			gre_subset["List"+" "+str(i)] = gre_index["List"+" "+str(i)]
		gre_flattened = []
		for sublist in gre_subset.values():
			for val in sublist:
				gre_flattened.append(val)

		print("\n")
		global test_quantity_wltype
		test_quantity_wltype = int(input(" \t 你想要考查多少单词？"))
		if test_quantity_wltype > len(gre_flattened):
			test_quantity_wltype = len(gre_flattened)
		word_list = rm.choices(population = gre_flattened, k = test_quantity_wltype)

		global indices
		indices = dict()
		for i in word_list:
			for j in range(0, len(i)):
				if j == 2:
					indices[i[j]] = [i[0], i[1]]

		global answers
		answers = dict()
		for i in indices.keys():
			answers[i] = gre["Answer"+" "+str(indices[i][0])][indices[i][1]]

		global scoreGained
		scoreGained = 0
		global scoreLost
		scoreLost = 0
		global wrongWordsList
		wrongWordsList = dict()
		print("\n")
		print("\t" + "请回答以下各个单词的中文含义是什么:" + "\n")
		for i in indices.keys():
			print("\t" + i + "\t")
			print("\n")
			usr_answer = input("\t" +"中文含义: \t")
			print("\n" + "\t\t\t\t\t\t\t标准答案是：" + str(answers[i]))
			print("\n")
			usr_scoring = input("回答是否正确（是/否）: ")
			if usr_scoring == "Y" or usr_scoring == "y" or usr_scoring == "是" or usr_scoring == "是 ":
				scoreGained += 1
			else:
				scoreLost += 1
				wrongWordsList[i] = answers[i]
			print("\n")
		print("\t" + "你回答了"+ str(test_quantity_wltype) +"个单词的含义；其中"+ str(scoreGained) +"个正确，"+ str(scoreLost) +"个错误。")
		print("\n")
		if scoreLost > 0:
			print("其中错误的单词为：")
			print("\n")
			for i in wrongWordsList.keys():
				print("\t\t\t\t" + str(i) + ", " + str(wrongWordsList[i]))
				print("\n")
		global test_quantity
		test_quantity = test_quantity_wltype
		review_wrong_words()
		continueTest = input("是否继续以该模式进行测试？")
		if continueTest == "Y" or continueTest == "y" or continueTest == "是":
			continueTest = True
		else:
			continueTest = False


def data_initialization():

	gre_original = pd.read_excel("GRE佛脚词.xlsx",header = 0)
	gre_original.index = np.arange(1, len(gre_original) + 1)

	global gre
	global gre_index
	gre = dict()
	gre_index = dict()
	for i in gre_original.columns.tolist():
		gre[i] = gre_original[i].dropna(axis = 0, how = "any").tolist()
		gre_index[i] = []
		for j in range(0, len(gre[i])):
			gre_index[i].append([int(i[-1]), j, gre[i][j]])

	global wrong_words_list_exist
	wrong_words_list_exist = os.path.isfile("GRE错题本.xlsx")
	global wb
	global wrong_words_list
	if wrong_words_list_exist == True:
		wb = load_workbook(filename = 'GRE错题本.xlsx', read_only=False)
		wrong_words_list = wb["错题本"]
	else:
		wb = Workbook()
		wrong_words_list = wb.active
		wrong_words_list.title = "错题本" 

def test_mode_choices():

	continueMode = True
	while continueMode:
		global usr_mode_choice
		usr_mode_choice = input("\t请输入你想测试的模式：（单元测试模式 or 综合测试模式）")
		if usr_mode_choice == "单元测试模式":
			test_initialization_lttype()
		elif usr_mode_choice == "综合测试模式":
			test_initialization_wltype()
		print("\n")
		continueMode = input("\t是否继续单词测试？")
		if continueMode == "Y" or continueMode == "y" or continueMode == "是":
			continueMode = True
		else:
			continueMode = False
	
def review_wrong_words():

	now = datetime.now()
	time = now.strftime("%m/%d/%Y %H:%M:%S")
	wrong_words_list.append(["".join(time)])

	if wrong_words_list_exist == False:
		for i in wrongWordsList.keys():
			wrong_words_list.append([i, wrongWordsList[i]])
	else:
		row = wrong_words_list.max_row + 1
		col = [1,2]
		for i, j in wrongWordsList.items():
			wrong_words_list.cell(row = row, column = col[0], value = i)
			wrong_words_list.cell(row = row, column = col[1], value = j)
			row += 1
		row = wrong_words_list.max_row + 1 
		wrong_words_list.cell(row = row, column = col[0], value = "")

	wb.save(r"GRE错题本.xlsx")

def main():

	print(format("Welcome to CUNY-Baruch Zicklin School of Business gre_original Self-testing Program", "*^116"))
	print(format("Designed by Sezaki Takahiro, version 1.2.2", "-^116"))
	print("\n")
	data_initialization()
	test_mode_choices()
	print("\n")
	print(format("谢谢使用，记得一键三连支持一下哦!!!", " ^100"))
main()

